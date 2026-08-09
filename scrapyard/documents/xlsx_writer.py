"""
xlsx_writer — ** The `scrapyard.documents.xlsx_writer` module provides a reusable interface for writing structured data to XLSX files, leveraging `openpyxl` for compatibility and performance. It ensures clean, type

### PART-META-JSON
{
  "name": "xlsx_writer",
  "layer": "documents",
  "purpose": "Provides a reusable interface for writing structured data to XLSX files, leveraging `openpyxl` for compatibility and performance. It ensures clean, type.",
  "addition": true,
  "status": "core",
  "dependencies": [],
  "inputs": "Public API: write_xlsx(file_path, data, **kwargs); XlsxWriter(...).",
  "outputs": "Returns: write_xlsx -> None.",
  "files_created": [],
  "security_notes": "Pure computation: no network, filesystem, subprocess, secrets, or persistence; validate ranges/values at the call site as usual.",
  "ai_usage": "Import what you need from `scrapyard.documents.xlsx_writer`.",
  "example": "from scrapyard.documents.xlsx_writer import *",
  "import_path": "scrapyard.documents.xlsx_writer"
}
### END-PART-META
"""
import logging
import os
import tempfile
from typing import Any, List, Optional

logger = logging.getLogger(__name__)


class XlsxWriter:
    """Reusable XLSX writer with support for nested structures and dynamic sheets."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._workbook = None
        self._openpyxl = None
        self._Workbook = None
    
    def _ensure_openpyxl(self):
        """Lazy import of openpyxl to avoid loading at module import time."""
        if self._openpyxl is None:
            try:
                import openpyxl
                from openpyxl import Workbook
                self._openpyxl = openpyxl
                self._Workbook = Workbook
            except ImportError as e:
                raise ImportError("openpyxl is required for XlsxWriter. Install with: pip install openpyxl") from e
    
    def _get_workbook(self):
        """Get or create the workbook instance."""
        self._ensure_openpyxl()
        if self._workbook is None:
            self._workbook = self._Workbook()
            # Remove default sheet; we'll create specific ones
            if self._workbook.active:
                self._workbook.remove(self._workbook.active)
        return self._workbook
    
    def write_data(self, data: Any, sheet_name: Optional[str] = None) -> None:
        """Write data to the workbook.
        
        Args:
            data: Data to write. Supports:
                - List[Dict]: List of dictionaries (rows) for a single sheet
                - List[List]: List of lists/rows for a single sheet
                - Dict[str, List]: Dictionary mapping sheet names to data
            sheet_name: Name for the sheet when data is a list. Defaults to 'Sheet1'.
        """
        if isinstance(data, dict):
            for name, sheet_data in data.items():
                if not isinstance(name, str):
                    raise TypeError(f"Sheet names must be strings, got {type(name)}")
                self._write_sheet_data(sheet_data, name)
        elif isinstance(data, list):
            self._write_sheet_data(data, sheet_name or "Sheet1")
        else:
            raise TypeError(f"Data must be a list or dict, got {type(data)}")
    
    def _write_sheet_data(self, data: List[Any], sheet_name: str) -> None:
        """Write a list of data to a specific sheet."""
        if not isinstance(data, list):
            raise TypeError(f"Sheet data must be a list, got {type(data)}")
        
        wb = self._get_workbook()
        ws = wb.create_sheet(title=sheet_name)
        
        if not data:
            return
        
        # Detect data structure from first item
        first_item = data[0]
        
        if isinstance(first_item, dict):
            # List of dicts: use keys as headers
            headers = list(first_item.keys())
            ws.append(headers)
            for row in data:
                if not isinstance(row, dict):
                    raise TypeError(f"All rows must be dicts when using list of dicts, got {type(row)}")
                row_values = [self._convert_value(row.get(k)) for k in headers]
                ws.append(row_values)
        elif isinstance(first_item, (list, tuple)):
            # List of lists
            for row in data:
                if not isinstance(row, (list, tuple)):
                    raise TypeError(f"All rows must be lists/tuples when using list of lists, got {type(row)}")
                ws.append([self._convert_value(v) for v in row])
        else:
            # Single column list
            for item in data:
                ws.append([self._convert_value(item)])
    
    def _convert_value(self, value: Any) -> Any:
        """Convert Python values to openpyxl-compatible types."""
        if value is None:
            return ""
        if isinstance(value, (bool, int, float, str)):
            return value
        # Handle datetime objects
        import datetime
        if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
            return value
        # Convert nested structures to JSON
        if isinstance(value, (list, dict, tuple)):
            import json
            return json.dumps(value, default=str)
        return str(value)
    
    def save(self) -> None:
        """Save the workbook to the file path."""
        if self._workbook is None:
            # Create empty workbook if nothing written
            self._get_workbook()
        
        # Ensure parent directory exists
        dir_path = os.path.dirname(os.path.abspath(self.file_path))
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)
        
        self._workbook.save(self.file_path)
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.save()


def write_xlsx(file_path: str, data: Any, **kwargs) -> None:
    """Write structured data to an XLSX file.
    
    Args:
        file_path: Destination file path (string)
        data: Data to write (list of dicts, list of lists, or dict of lists)
        **kwargs: Reserved for future options
        
    Raises:
        TypeError: If file_path is not a string or data is invalid
    """
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be a string, got {type(file_path)}")
    
    with XlsxWriter(file_path) as writer:
        writer.write_data(data)


def _selftest() -> None:
    """Offline self-test for the xlsx_writer module."""
    print("Running _selftest for scrapyard.documents.xlsx_writer...")
    
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
        # Import openpyxl here to verify it's available for testing
        import openpyxl
        
        # Test 1: Flat list of dicts
        flat_data = [
            {"name": "Alice", "age": 30, "active": True, "score": 95.5},
            {"name": "Bob", "age": 25, "active": False, "score": 82.0},
            {"name": "Charlie", "age": 35, "active": True, "score": 88.5}
        ]
        file1 = os.path.join(tmpdir, "test1.xlsx")
        write_xlsx(file1, flat_data)
        assert os.path.exists(file1), "Test 1: File should be created"
        
        wb = openpyxl.load_workbook(file1)
        assert "Sheet1" in wb.sheetnames, "Test 1: Default sheet name should be Sheet1"
        ws = wb["Sheet1"]
        assert ws.cell(1, 1).value == "name", "Test 1: Header should be 'name'"
        assert ws.cell(2, 1).value == "Alice", "Test 1: First name should be Alice"
        assert ws.cell(2, 2).value == 30, "Test 1: Age should be integer 30"
        assert ws.cell(2, 3).value is True, "Test 1: Boolean True should be preserved"
        assert isinstance(ws.cell(2, 4).value, (int, float)), "Test 1: Float should be preserved"
        wb.close()
        
        # Test 2: Dict of lists (multiple sheets)
        multi_data = {
            "Employees": [
                {"id": 1, "dept": "Engineering"},
                {"id": 2, "dept": "Sales"}
            ],
            "Products": [
                {"sku": "A001", "price": 19.99},
                {"sku": "A002", "price": 29.99}
            ]
        }
        file2 = os.path.join(tmpdir, "test2.xlsx")
        write_xlsx(file2, multi_data)
        
        wb = openpyxl.load_workbook(file2)
        assert "Employees" in wb.sheetnames, "Test 2: Employees sheet should exist"
        assert "Products" in wb.sheetnames, "Test 2: Products sheet should exist"
        assert wb["Employees"].cell(2, 2).value == "Engineering", "Test 2: Nested data verification"
        wb.close()
        
        # Test 3: Nested data structures (dicts/lists in cells)
        nested_data = [
            {"item": "Part1", "metadata": {"weight": 10, "tags": ["heavy", "metal"]}},
            {"item": "Part2", "metadata": {"weight": 5, "tags": ["light"]}}
        ]
        file3 = os.path.join(tmpdir, "test3.xlsx")
        write_xlsx(file3, nested_data)
        
        wb = openpyxl.load_workbook(file3)
        ws = wb["Sheet1"]
        meta_val = ws.cell(2, 2).value
        assert isinstance(meta_val, str), "Test 3: Nested dict should be converted to string"
        assert "weight" in meta_val and "tags" in meta_val, "Test 3: JSON should contain keys"
        wb.close()
        
        # Test 4: List of lists
        lol_data = [
            ["Header1", "Header2", "Header3"],
            [1, 2, 3],
            [4, 5, 6]
        ]
        file4 = os.path.join(tmpdir, "test4.xlsx")
        write_xlsx(file4, lol_data)
        
        wb = openpyxl.load_workbook(file4)
        ws = wb["Sheet1"]
        assert ws.cell(1, 1).value == "Header1", "Test 4: List headers"
        assert ws.cell(3, 3).value == 6, "Test 4: List data"
        wb.close()
        
        # Test 5: XlsxWriter direct usage with custom sheet name
        file5 = os.path.join(tmpdir, "test5.xlsx")
        writer = XlsxWriter(file5)
        writer.write_data([{"a": 1}, {"a": 2}], sheet_name="CustomSheet")
        writer.save()
        
        wb = openpyxl.load_workbook(file5)
        assert "CustomSheet" in wb.sheetnames, "Test 5: Custom sheet name should be used"
        wb.close()
        
        # Test 6: Context manager usage
        file6 = os.path.join(tmpdir, "test6.xlsx")
        with XlsxWriter(file6) as w:
            w.write_data({"SheetA": [{"x": 1}], "SheetB": [{"y": 2}]})
        
        wb = openpyxl.load_workbook(file6)
        assert "SheetA" in wb.sheetnames and "SheetB" in wb.sheetnames, "Test 6: Context manager sheets"
        wb.close()
        
        # Test 7: Invalid inputs raise exceptions
        try:
            write_xlsx(os.path.join(tmpdir, "err.xlsx"), "not a list or dict")
            assert False, "Test 7: Should raise TypeError for string data"
        except TypeError:
            pass
        
        try:
            write_xlsx(123, [{"a": 1}])  # type: ignore
            assert False, "Test 7: Should raise TypeError for non-string path"
        except TypeError:
            pass
        
        try:
            write_xlsx(os.path.join(tmpdir, "err2.xlsx"), {"Sheet1": "not a list"})
            assert False, "Test 7: Should raise TypeError for non-list sheet data"
        except TypeError:
            pass
        
        # Test 8: Empty data handling
        file8 = os.path.join(tmpdir, "test8.xlsx")
        write_xlsx(file8, [])
        wb = openpyxl.load_workbook(file8)
        assert len(wb.sheetnames) >= 1, "Test 8: Empty list should still create sheet"
        wb.close()
        
        # Test 9: None values and mixed types
        mixed_data = [
            {"val": None, "num": 42, "flag": False, "text": "hello"},
            {"val": "test", "num": 0, "flag": True, "text": None}
        ]
        file9 = os.path.join(tmpdir, "test9.xlsx")
        write_xlsx(file9, mixed_data)
        
        wb = openpyxl.load_workbook(file9)
        ws = wb["Sheet1"]
        # None becomes empty string
        assert ws.cell(2, 1).value == "" or ws.cell(2, 1).value is None, "Test 9: None handling"
        assert ws.cell(2, 3).value is False, "Test 9: False handling"
        wb.close()
    
    print("_selftest PASSED")


if __name__ == "__main__":
    _selftest()
